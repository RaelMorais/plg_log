import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.drawing.image import Image
import os

# Obtenha o diretório atual do script
diretorio_script = os.path.dirname(os.path.abspath(__file__))

caminho_relatorio_movimentacao = os.path.join(diretorio_script, 'Relatorios', 'CSV', 'relatorio_movimentacao.csv')
caminho_relatorio_produto = os.path.join(diretorio_script, 'Relatorios', 'CSV', 'relatorio_produto.csv')
diretorio_Grafico = os.path.join(diretorio_script, 'Relatorios', 'Graficos')
diretorio_Relatorio = os.path.join(diretorio_script, 'Relatorios')

# Lê o arquivo CSV com o encoding 'utf-8' e ignore os caracteres inválidos
try:
    df_movimentacao = pd.read_csv(caminho_relatorio_movimentacao, encoding='utf-8')
    df_produto = pd.read_csv(caminho_relatorio_produto, encoding='utf-8')
except UnicodeDecodeError:
    df_movimentacao = pd.read_csv(caminho_relatorio_movimentacao, encoding='ISO-8859-1')
    df_produto = pd.read_csv(caminho_relatorio_produto, encoding='ISO-8859-1')

df_movimentacao = df_movimentacao.rename(columns={'Responsavel': 'Responsavel', 'Codigo Pallet': 'Codigo'})

df_movimentacao = df_movimentacao.drop(['Responsavel'], axis=1)

dfs_por_codigo = {codigo: grupo for codigo, grupo in df_movimentacao.groupby('Codigo')}

codigo_para_nome = dict(zip(df_produto['Codigo'], df_produto['Nome']))

totais_entrada = []
totais_saida = []

nomes_produtos = []

for codigo, df_codigo in dfs_por_codigo.items():

    df_entrada = df_codigo[df_codigo['Movimentacao'] == 1]
    df_saida = df_codigo[df_codigo['Movimentacao'] == 2]

    colunas_pallet = ['Pallet1', 'Pallet2', 'Pallet3', 'Pallet4', 'Pallet5', 'Pallet6']
    total_entrada = df_entrada[colunas_pallet].sum(axis=1).sum()
    total_saida = df_saida[colunas_pallet].sum(axis=1).sum()

    totais_entrada.append(total_entrada)
    totais_saida.append(total_saida)

    nomes_produtos.append(codigo_para_nome.get(codigo, 'Desconhecido'))

bar_width = 0.35
index = np.arange(len(dfs_por_codigo))

plt.figure(figsize=(10, 6))
plt.bar(index, totais_entrada, label='Entrada', width=bar_width)
plt.bar(index + bar_width, totais_saida, label='Saida', width=bar_width)
plt.xlabel('Codigo')
plt.ylabel('Total')
plt.title('Comparacao de Entrada e Saida por Codigo')
plt.xticks(index + bar_width / 2, nomes_produtos)
plt.legend()
plt.xticks(rotation=90)
plt.tight_layout()

# Obtenha a data atual no formato 'ddmmyyyy'
data_atual = datetime.now().strftime('%d%m%Y')

# Salve o grafico em uma imagem na pasta de gráficos
nome_arquivo_grafico = os.path.join(diretorio_Grafico, f'grafico_{data_atual}.png')
plt.savefig(nome_arquivo_grafico)

# Salve o DataFrame com datas em um arquivo 'relatorio_data.xlsx' na pasta de relatórios
nome_arquivo_excel = os.path.join(diretorio_Relatorio, f'relatorio_{data_atual}.xlsx')
with pd.ExcelWriter(nome_arquivo_excel, engine='openpyxl') as writer:
    df_movimentacao.to_excel(writer, sheet_name=f'movimentacao', index=False)
    workbook = writer.book
    worksheet = writer.sheets[f'movimentacao']

    for col in worksheet.iter_cols(min_col=1, max_col=1, min_row=2, max_row=len(df_movimentacao) + 1):
        for cell in col:
            cell.number_format = '0'
    
    # Adicione o grafico a planilha de movimentacao
    img = Image(nome_arquivo_grafico)
    img.width = 600
    img.height = 400
    worksheet.add_image(img, 'E2')

    df_produto.to_excel(writer, sheet_name=f'produto', index=False)
    worksheet = writer.sheets[f'produto']
    
    for col in worksheet.iter_cols(min_col=1, max_col=1, min_row=2, max_row=len(df_produto) + 1):
        for cell in col:
            cell.number_format = '0'
    
print("Relatorio gerado com sucesso!")